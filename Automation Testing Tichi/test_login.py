from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
import time
workbook=load_workbook(r"C:\Users\Manothakin\Downloads\sample data tichi.xlsx")
sheet=workbook.active
driver=webdriver.Chrome()
driver.maximize_window()
wait=WebDriverWait(driver,20)
rows=sheet.max_row
for i in range(2,rows+1):
    email=str(sheet.cell(row=i,column=1).value)
    password=str(sheet.cell(row=i,column=2).value)
    print(f"\nTesting Row {i}")
    print(email)
    driver.delete_all_cookies()
    driver.get("https://tichi-app-webapp-stage.web.app/login")
    time.sleep(3)
    try:
        email_box = wait.until(
            EC.visibility_of_element_located((By.ID,"email"))
        )
        email_box.clear()
        email_box.send_keys(email)
        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH,"/html/body/div[1]/div[2]/div/form/button")
            )
        ).click()
        password_box = wait.until(
            EC.visibility_of_element_located((By.ID,"password"))
        )
        password_box.clear()
        password_box.send_keys(password)
        wait.until(
            EC.element_to_be_clickable(
                (By.XPATH,"/html/body/div[1]/div[2]/div/form/button")
            )
        ).click()
        time.sleep(5)
        if "home" in driver.current_url:
            print(email,"--> Valid User")
        else:
            print(email,"--> Invalid User")
    except TimeoutException:
        print("Element Not Found")
        print("Current URL :",driver.current_url)
    except Exception as e:
        print("Unexpected Error :",e)
driver.quit()